"""Assemble the final long-format Metal_Intensity table and write both output
artifacts: technologies_mi_all_years.xlsx and ampl_files/Material_intensity.dat.

Every technology in the Mapping sheet (Material_intensities_energyscope.xlsx) is
recomputed on every run: techs with a real mapping_type get their values from the
literature source data (currently only the ~35 electricity/fuel-cell ones -- see
canonical.py), techs marked not_mapped get blank cells. Anything not in the
Mapping sheet at all is carried through unchanged from the current
technologies_mi_all_years.xlsx.

Performance note: rows are written with openpyxl's write_only Workbook, which streams
straight to disk instead of building an in-memory cell index -- the normal-mode
per-cell-style approach used to hand-edit this file degrades badly past ~100k rows
(confirmed: 4h15 to write 184k rows). write_only avoids that entirely.
"""
import time
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill

from . import canonical, groups, sources
from .aggregate import (VEHICLE_POWERTRAINS, PUBLIC_TRANSIT_POWERTRAINS, YEARS, compute_all,
                         compute_vehicle_intensities_bieuville, compute_vehicle_intensities_public_transit)
from .mapping import load_mapping

_PROJ_ROOT = Path(__file__).resolve().parents[2]  # .../projects/critical_materials
CURRENT_XLSX = _PROJ_ROOT / 'excel_files' / 'technologies_mi_all_years.xlsx'
OUT_XLSX = CURRENT_XLSX
OUT_DAT_NAME = 'Material_intensity'

FONT_A = Font(name='Arial', size=12)
_GROUP_FILLS = {
    group: PatternFill(start_color=f'FF{hexcolor}', end_color=f'FF{hexcolor}', fill_type='solid')
    for group, hexcolor in groups.GROUP_COLORS.items()
}

def load_material_output_order(path=sources.SOURCE_XLSX):
    """Output column order for materials -- the short codes from the
    'Materials' sheet, in row order. Add a material by adding a row there
    (see sources.load_materials()); nothing here needs to change."""
    return list(sources.load_materials(path).values())


MATERIAL_OUTPUT_ORDER = load_material_output_order()


def _column_fill(tech):
    return _GROUP_FILLS.get(groups.categorize(tech))


def _read_existing_unmapped_rows(mapped_scope):
    """Read technologies_mi_all_years.xlsx's Metal_Intensity sheet and return every
    row (as a 7-tuple) whose tech isn't in `mapped_scope` (i.e. not in the Mapping
    sheet at all), in their original order -- left completely untouched."""
    wb = openpyxl.load_workbook(CURRENT_XLSX, data_only=False)
    ws = wb['Metal_Intensity']
    rows = []
    for r in range(2, ws.max_row + 1):
        row = tuple(ws.cell(row=r, column=c).value for c in range(1, 8))
        if row[2] not in mapped_scope:
            rows.append(row)
    return rows


def _mapped_rows(mapping, intensities, vehicle_source='watari'):
    """Long-format rows for every technology in the Mapping sheet (in scope), in
    tech -> MATERIAL_OUTPUT_ORDER -> YEARS order. not_mapped techs get blank
    (None) Values, which create_dat_file_from_excel then skips entirely."""
    rows = []
    for tech, row in mapping.iterrows():
        df = intensities[tech]
        is_vehicle = bool(row['subtechs']) and set(row['subtechs']) <= VEHICLE_POWERTRAINS
        is_fcv = is_vehicle and row['subtechs'][0] == 'FCV'
        is_public = is_vehicle and row['subtechs'][0] in PUBLIC_TRANSIT_POWERTRAINS
        unit = 't/(pkm/h)' if is_vehicle else 't/GW'
        if row['mapping_type'] == 'not_mapped':
            comment = f"[not_mapped] {row['notes']}".strip()
        else:
            subtechs = ','.join(row['subtechs'])
            if not is_vehicle:
                source = 'Bieuville et al. 2025 (MI_Energy)'
            elif is_public:
                source = 'Månberger & Stenqvist 2018 (MI_Vehicles_Public + MS_Battery_Motor_LDV)'
            elif is_fcv or vehicle_source == 'watari':
                # FCV always falls back to MI_Vehicles regardless of vehicle_source (Bieuville doesn't cover it)
                source = 'Watari et al. 2019 / Fishman et al. 2018 (MI_Vehicles)'
            else:
                source = 'Bieuville et al. 2025 (MI_Vehicles_Bieuville_Clean + MS_Battery_Motor_LDV)'
            comment = (f"[{row['confidence']}] {source}; "
                       f"mapping: {row['mapping_type']} <- {subtechs}. See the Mapping sheet.")
        for material in MATERIAL_OUTPUT_ORDER:
            for year in YEARS:
                raw_value = df.loc[material, year]
                value = None if pd.isna(raw_value) else float(raw_value)
                rows.append(('material_intensity', year, tech, material, value, unit, comment))
    return rows


def _vehicle_calc_detail_rows(mapping, mi_vehicles, ref_size, vehicle_intensities_g):
    """One row per (tech, year, material) for every vehicle-mapped technology
    (mapping_type='direct', subtechs one of VEHICLE_POWERTRAINS): the
    intermediate g/vehicle value, the ref_size used, and the final
    material_intensity -- so the g/vehicle -> t/(pkm/h) conversion (which never
    appears in the Excel elsewhere, see aggregate.compute_tech_intensity) is
    auditable. Public-transit powertrains (ICEV_PUBLIC/HEV_PUBLIC/EV_PUBLIC) and,
    under vehicle_source='bieuville', private ones too, come from
    vehicle_intensities_g and vary by year; anything else falls back to the flat
    MI_Vehicles value repeated every year."""
    rows = []
    for tech, row in mapping.iterrows():
        if not (row['mapping_type'] == 'direct' and len(row['subtechs']) == 1
                and row['subtechs'][0] in VEHICLE_POWERTRAINS):
            continue
        powertrain = row['subtechs'][0]
        family = canonical.family_of(tech)
        from_table = powertrain in vehicle_intensities_g
        for material in MATERIAL_OUTPUT_ORDER:
            if not from_table and material not in mi_vehicles.index:
                continue  # not tracked in MI_Vehicles (e.g. electrolyzer-only materials like Ti/Ir/La)
            for year in YEARS:
                if from_table:
                    mi_g = float(vehicle_intensities_g[powertrain].loc[material, year])
                else:
                    mi_g = float(mi_vehicles.loc[material, powertrain])
                r = ref_size.get((year, family))
                total = None if r is None else mi_g * 1e-6 / r
                rows.append((tech, year, material, powertrain, mi_g, r, total))
    return rows


def _write_xlsx(all_rows, vehicle_detail_rows, path=OUT_XLSX):
    wb = Workbook(write_only=True)

    ws = wb.create_sheet('Metal_Intensity')
    header = ['Parameter', 'index0', 'index1', 'index2', 'Value', 'Unit', 'Comment']
    header_cells = [WriteOnlyCell(ws, value=h) for h in header]
    for cell in header_cells:
        cell.font = Font(name='Calibri', size=11, bold=True)
    ws.append(header_cells)

    for row in all_rows:
        param_cell = WriteOnlyCell(ws, value=row[0])
        param_cell.font = FONT_A
        tech_cell = WriteOnlyCell(ws, value=row[2])
        fill = _column_fill(row[2])
        if fill is not None:
            tech_cell.fill = fill
        ws.append([param_cell, row[1], tech_cell, row[3], row[4], row[5], row[6]])

    legend = wb.create_sheet('Legend')
    legend.append([None, None, 'Legend'])
    legend.append([None, None])
    for group, _keywords in groups.CATEGORY_RULES:
        swatch = WriteOnlyCell(legend, value=None)
        swatch.fill = _GROUP_FILLS[group]
        legend.append([None, swatch, groups.GROUP_LABELS[group]])

    detail = wb.create_sheet('Vehicle_Calc_Detail')
    detail_header = ['tech', 'year', 'material', 'powertrain', 'mi_g_per_vehicle',
                      'ref_size', 'material_intensity_t_per_pkmh']
    detail_header_cells = [WriteOnlyCell(detail, value=h) for h in detail_header]
    for cell in detail_header_cells:
        cell.font = Font(name='Calibri', size=11, bold=True)
    detail.append(detail_header_cells)
    for row in vehicle_detail_rows:
        detail.append(list(row))

    wb.save(path)


def create_dat_file_from_excel(df, file_name, out_dir=None, materials=MATERIAL_OUTPUT_ORDER):
    """Adapted from `New parameters and constraints.ipynb` (cell 5) -- writes
    ampl_files/{file_name}.dat from a long-format DataFrame
    (Parameter/index0/index1/index2/Value/Unit/Comment).

    The `set MATERIALS := ...` line is derived from `materials` (MATERIAL_OUTPUT_ORDER
    by default) instead of being duplicated as a separate hardcoded string. Both that
    and MATERIAL_OUTPUT_ORDER itself come from the 'Materials' sheet (sources.py) --
    adding a material there is enough, nothing in this code needs to change."""
    out_dir = out_dir or (_PROJ_ROOT / 'ampl_files')
    out_path = Path(out_dir) / f'{file_name}.dat'
    with open(out_path, 'w', encoding='utf-8', newline='\n') as f:
        f.write("data;\n\n")
        f.write(f"set MATERIALS := {' '.join(materials)} ;\n \n")
        for _, row in df.iterrows():
            value = row['Value']
            if pd.isna(value):
                continue  # skip missing values, params already default to 0
            param_name = row['Parameter']
            index0 = row['index0']
            index1 = row['index1']
            index2 = row['index2']
            unit = '-' if pd.isna(row.get('Unit')) else str(row.get('Unit'))
            comment = '' if pd.isna(row.get('Comment')) else str(row.get('Comment'))
            if pd.isna(index1) and pd.isna(index2):
                f.write(f"let {param_name}['{index0}'] := {value} ; # [{unit}] {comment}\n")
            elif pd.isna(index2):
                f.write(f"let {param_name}['{index0}','{index1}'] := {value} ; # [{unit}] {comment}\n")
            else:
                f.write(f"let {param_name}['{index0}','{index1}','{index2}'] := {value} ; # [{unit}] {comment}\n")
    return out_path


def build(scenario='baseline', vehicle_source='watari', write_dat=True, write_xlsx=True):
    """vehicle_source: 'watari' (default) is the original flat MI_Vehicles-based
    computation. 'bieuville' uses MI_Vehicles_Bieuville_Clean +
    MS_Battery_Motor_LDV instead (see aggregate.compute_vehicle_intensities_bieuville).
    Either way this writes to the same technologies_mi_all_years.xlsx/
    Material_intensity.dat filenames -- rerunning with a different
    vehicle_source overwrites them, it doesn't keep both around. To compare
    the two, build+run_pathway_materials with one, save/rename the results,
    then build+run again with the other."""
    t0 = time.time()
    mapping = load_mapping()

    # A not_mapped tech is always safe to include (it only ever produces blank
    # cells, which create_dat_file_from_excel skips) -- but a tech claiming *real*
    # data has to already be declared in QC_data.dat, or AMPL chokes on an
    # out-of-set subscript when the .dat file is loaded. Mapping rows for a
    # not-yet-modelled tech stay in the Mapping sheet for later, just excluded
    # from output until it's added to the model.
    canonical_techs = set(canonical.all_target_techs())
    claims_real_data = mapping['mapping_type'] != 'not_mapped'
    not_yet_modeled = set(mapping.index[claims_real_data]) - canonical_techs
    if not_yet_modeled:
        print(f"[build_table] skipping (not yet in QC_data.dat): {sorted(not_yet_modeled)}")
    mapped_scope = set(mapping.index) - not_yet_modeled
    mapping = mapping.loc[sorted(mapped_scope)]

    intensities = compute_all(scenario=scenario, vehicle_source=vehicle_source)
    print(f"[build_table] computed {len(intensities)} tech intensities in {time.time()-t0:.1f}s")

    unmapped_existing_rows = _read_existing_unmapped_rows(mapped_scope)
    print(f"[build_table] kept {len(unmapped_existing_rows)} existing rows for techs outside the Mapping sheet in {time.time()-t0:.1f}s")

    mapped_rows = _mapped_rows(mapping, intensities, vehicle_source=vehicle_source)
    print(f"[build_table] built {len(mapped_rows)} rows for the Mapping sheet's {len(mapping)} technologies in {time.time()-t0:.1f}s")

    all_rows = unmapped_existing_rows + mapped_rows

    out_xlsx = CURRENT_XLSX
    out_dat_name = OUT_DAT_NAME

    if write_xlsx:
        mi_vehicles = sources.load_mi_vehicles()
        ref_size = canonical.load_ref_size()
        materials_idx = pd.concat([sources.load_mi_energy(), mi_vehicles, sources.load_mi_h2()], axis=1).index
        vehicle_intensities_g = compute_vehicle_intensities_public_transit(materials_idx)
        if vehicle_source == 'bieuville':
            vehicle_intensities_g.update(compute_vehicle_intensities_bieuville(materials_idx))
        vehicle_detail_rows = _vehicle_calc_detail_rows(mapping, mi_vehicles, ref_size,
                                                          vehicle_intensities_g=vehicle_intensities_g)
        _write_xlsx(all_rows, vehicle_detail_rows, path=out_xlsx)
        print(f"[build_table] wrote {out_xlsx.name} ({len(all_rows)} rows, "
              f"{len(vehicle_detail_rows)} vehicle-detail rows) in {time.time()-t0:.1f}s")

    if write_dat:
        df = pd.DataFrame(all_rows, columns=['Parameter', 'index0', 'index1', 'index2', 'Value', 'Unit', 'Comment'])
        out_path = create_dat_file_from_excel(df, out_dat_name)
        print(f"[build_table] wrote {out_path.name} in {time.time()-t0:.1f}s")

    print(f"[build_table] total: {time.time()-t0:.1f}s")
    return all_rows


if __name__ == '__main__':
    build()
