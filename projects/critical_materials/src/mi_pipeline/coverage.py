"""Coverage report: which EnergyScope technologies actually have usable material-
intensity data, vs. placeholder zeros, vs. genuinely unmapped.
"""
import openpyxl
import pandas as pd

from . import canonical
from .mapping import load_mapping, MAPPING_XLSX

STATUS_ORDER = ['not_mapped', 'not_yet_modeled', 'placeholder_zero', 'integrated']


def build_report(mapping, intensities):
    """DataFrame indexed by energyscope_tech: mapping_type, confidence, status.

    status is:
      - 'not_mapped'       if mapping_type == 'not_mapped'
      - 'not_yet_modeled'  if the tech isn't (yet) in QC_data.dat -- mapped here for
                           later, but excluded from the actual output files
      - 'placeholder_zero' if mapped but every material/year value is 0
      - 'integrated'       otherwise (at least one nonzero value)
    """
    canonical_techs = set(canonical.all_target_techs())
    rows = []
    for tech, row in mapping.iterrows():
        if tech not in canonical_techs:
            status = 'not_yet_modeled'
        elif row['mapping_type'] == 'not_mapped':
            status = 'not_mapped'
        elif (intensities[tech] == 0).all().all():
            status = 'placeholder_zero'
        else:
            status = 'integrated'
        rows.append({
            'energyscope_tech': tech,
            'mapping_type': row['mapping_type'],
            'confidence': row['confidence'],
            'status': status,
        })
    return pd.DataFrame(rows).set_index('energyscope_tech')


def print_report(report):
    counts = report['status'].value_counts().reindex(STATUS_ORDER, fill_value=0)
    print("Coverage report:")
    for status in STATUS_ORDER:
        print(f"  {status:17s}: {counts[status]}")
    for status in STATUS_ORDER:
        techs = sorted(report.index[report['status'] == status])
        if techs:
            print(f"\n{status} ({len(techs)}):")
            for t in techs:
                print(f"  - {t}")


def write_status_back(report, path=MAPPING_XLSX):
    """Refresh the (auto-computed) 'status' column in tech_mapping.xlsx's Mapping
    sheet in place, so it's visible when reviewing the file in Excel."""
    wb = openpyxl.load_workbook(path)
    ws = wb['Mapping']
    header = [c.value for c in ws[1]]
    tech_col = header.index('energyscope_tech') + 1
    status_col = header.index('status') + 1
    for r in range(2, ws.max_row + 1):
        tech = ws.cell(row=r, column=tech_col).value
        if tech in report.index:
            ws.cell(row=r, column=status_col, value=report.loc[tech, 'status'])
    wb.save(path)


if __name__ == '__main__':
    from .aggregate import compute_all
    mapping = load_mapping()
    intensities = compute_all()
    report = build_report(mapping, intensities)
    print_report(report)
    write_status_back(report)
    print(f"\nStatus column refreshed in {MAPPING_XLSX.name}")
